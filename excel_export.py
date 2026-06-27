from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


WRAP_COLUMNS = {"原文片段", "建議文字", "批註內容", "原因"}


def build_issues_excel(rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "校對建議"

    if not rows:
        worksheet.append(["沒有校對建議"])
    else:
        headers = list(rows[0].keys())
        worksheet.append(headers)

        for row in rows:
            worksheet.append([row.get(header, "") for header in headers])

        header_fill = PatternFill("solid", fgColor="D9EAF7")

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column_index, header in enumerate(headers, start=1):
            column_letter = get_column_letter(column_index)

            width = 14
            if header in {"issue_id", "原文片段", "建議文字", "批註內容", "原因"}:
                width = 28
            elif header in {"位置", "類別代碼", "類別", "分級"}:
                width = 16

            worksheet.column_dimensions[column_letter].width = width

            if header in WRAP_COLUMNS:
                for cell in worksheet[column_letter]:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()